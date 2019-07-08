# -*- coding: utf-8 -*-
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
wb = Workbook()    #创建文件对象
ws=wb.active

# ws = wb.active     #获取第一个sheet
filename = 'D:\\011111111111111111111111\\axiba.xlsx'


test_book = xlsxwriter.Workbook(filename)
worksheet = test_book.add_worksheet('what')
bold = test_book.add_format({'bold': True})


expenses = [
    ['Rent', 1000],
    ['Gas',   100],
    ['Food',  300],
    ['Gym',    50],
]

# 定义起始的行列 会在这个基础上 行列各加一 作为初始行列
row = 0
col = 0

for i in range(len(expenses)):
    data={}
    for col in range(len(expenses[i])):
        data[col+1]=expenses[i][col]
    ws.append(data)
    # worksheet.write(row, col, item)
    # worksheet.write(row, col+1, cost)
    # worksheet.insert_image(row, col+2,'D:\\tmp\\wj\\location.png')
    # ws.add_image()
    ws.add_image('D:\\tmp\\wj\\location.png')
    # row += 1
wb.save(filename)
# worksheet.write(row, col, '=sum(B0:B4)')
# test_book.close()